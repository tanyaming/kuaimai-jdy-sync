#!/usr/bin/env python3
"""
快麦后台导出 → 拼多多平台订单号(tid)映射提取器

用法:
  python3 scripts/extract_tid_map.py <导出文件.xlsx> [...更多文件] -o pdd_tid_map.json

功能:
  读取快麦ERP「交易订单导出」xlsx，提取所有平台的 (系统订单号 sid -> 平台订单号 tid) 映射，
  默认只输出拼多多，可加 --all 输出全部平台。

输出: JSON 文件 { sid: tid }，供服务器容器内 update_tid.js 补录简道云。
"""
import openpyxl, warnings, json, sys, argparse

warnings.filterwarnings('ignore')

def extract(path, only_pdd=True):
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue
        hdr = rows[1]
        # 找到列索引
        idx = {}
        for i, h in enumerate(hdr):
            if h == '系统订单号':
                idx['sid'] = i
            elif h == '平台订单号':
                idx['tid'] = i
            elif h == '平台':
                idx['plat'] = i
        if 'sid' not in idx or 'tid' not in idx:
            continue
        for r in rows[2:]:
            plat = r[idx['plat']] if 'plat' in idx and idx['plat'] < len(r) else None
            if only_pdd and plat != '拼多多':
                continue
            sid_raw = r[idx['sid']]
            tid_raw = r[idx['tid']]
            if sid_raw is None:
                continue
            # 系统订单号：若 openpyxl 读成 float（如 1.23e15）需转整数再转字符串；
            # 若本身就是字符串（含末尾 0），直接保留，切勿 rstrip('0') 否则丢末尾 0。
            if isinstance(sid_raw, float):
                sid = str(int(sid_raw))
            else:
                sid = str(sid_raw).strip()
            tid = str(tid_raw).strip() if tid_raw not in (None, '') else ''
            if sid and sid not in ('None', ''):
                # 保留首次出现的非空 tid
                if tid and (sid not in out or not out[sid]):
                    out[sid] = tid
                elif sid not in out:
                    out[sid] = tid
    return out

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('files', nargs='+', help='xlsx 导出文件')
    ap.add_argument('-o', '--output', default='pdd_tid_map.json')
    ap.add_argument('--all', action='store_true', help='提取全部平台（默认只拼多多）')
    args = ap.parse_args()

    merged = {}
    for f in args.files:
        m = extract(f, only_pdd=not args.all)
        for sid, tid in m.items():
            if tid and (sid not in merged or not merged[sid]):
                merged[sid] = tid
            elif sid not in merged:
                merged[sid] = tid

    empty = [k for k, v in merged.items() if not v]
    with open(args.output, 'w', encoding='utf-8') as fp:
        json.dump(merged, fp, ensure_ascii=False, indent=2)

    print(f'共 {len(merged)} 个 sid，其中 tid 为空 {len(empty)} 个')
    print(f'已写入 {args.output}')

if __name__ == '__main__':
    main()
